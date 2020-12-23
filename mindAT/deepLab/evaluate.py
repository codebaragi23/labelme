"""Evaluate a DeepLab v3 model."""

from __future__ import absolute_import
from __future__ import division
from __future__ import print_function

import os
import tensorflow as tf
import cv2
import glob
import numpy as np
from openpyxl import Workbook, load_workbook
import time
import scipy.ndimage
from PIL import Image
import imageio

from mindAT.deepLab import deeplab_model
from mindAT.deepLab.utils import preprocessing, dataset_util
from mindAT.deepLab.config import *

DATA_TYPE = tf.uint8


def parse_record(raw_record):
	"""Parse PASCAL image and label from a tf record."""
	keys_to_features = {
		'row':
			tf.FixedLenFeature((), tf.int64),
		'col':
			tf.FixedLenFeature((), tf.int64),
		'image_raw':
			tf.FixedLenFeature((), tf.string, default_value=''),
		'label_raw':
			tf.FixedLenFeature((), tf.string, default_value=''),
		'row_off':
			tf.FixedLenFeature((), tf.int64),
		'col_off':
			tf.FixedLenFeature((), tf.int64),
		'filename':
			tf.FixedLenFeature((), tf.string, default_value='')
	}

	parsed = tf.parse_single_example(raw_record, keys_to_features)

	image = tf.decode_raw(parsed['image_raw'], DATA_TYPE)
	image = tf.reshape(image, [PATCH_SIZE, PATCH_SIZE, DATA_DEPTH])
	image = image[..., :DEPTH]
	image = tf.to_float(tf.image.convert_image_dtype(image, dtype=DATA_TYPE))

	label = tf.decode_raw(parsed['label_raw'], tf.uint8)
	label = tf.reshape(label, [PATCH_SIZE, PATCH_SIZE, 1])
	label = tf.to_int32(tf.image.convert_image_dtype(label, dtype=tf.uint8))

	row = tf.cast(parsed['row'], tf.int32)
	col = tf.cast(parsed['col'], tf.int32)
	row_off = tf.cast(parsed['row_off'], tf.int32)
	col_off = tf.cast(parsed['col_off'], tf.int32)
	filename = tf.cast(parsed['filename'], tf.string)

	image = preprocessing.mean_image_subtraction(image)

	return image, label, row, col, row_off, col_off, filename


def input_fn(tfrecord_file, batch_size, num_epochs=1):
	dataset = tf.data.Dataset.from_tensor_slices([tfrecord_file])
	dataset = dataset.flat_map(tf.data.TFRecordDataset)

	dataset = dataset.map(parse_record)
	dataset = dataset.prefetch(batch_size)

	# We call repeat after shuffling, rather than before, to prevent separate
	# epochs from blending together.
	dataset = dataset.repeat(num_epochs)
	dataset = dataset.batch(batch_size)

	iterator = dataset.make_one_shot_iterator()
	images, labels, row, col, row_offs, col_offs, filenames = iterator.get_next()

	return images, labels, row, col, row_offs, col_offs, filenames


def calc_accuracy(gt_file, result_file, class_num):
	print("gt:", gt_file, ", result:", result_file)
	gt = dataset_util.load_label(gt_file)
	result = dataset_util.load_label(result_file)
	#result = cv2.imread(result_file)
	#result = result[:, :, 0]

	result[result > (class_num - 1)] = 0
	gt[gt > (class_num - 1)] = 0

	# st = time.time()
	count_table = np.zeros([class_num, class_num], dtype=np.int)
	for x in range(gt.shape[0]):
		for y in range(gt.shape[1]):
			count_table[gt[x][y]][result[x][y]] += 1
	for a, b in zip(gt, result):
		count_table[a][b] += 1
	# print('%.2f' % (time.time() - st))
	# print(count_table)

	return count_table

def save_result(filename, vote, result_path):
	#vote[:, :, 0] = 0
	vote = np.argmax(vote, axis=-1)
	vote = np.uint8(vote)

	result_file = os.path.join(result_path, filename)
	visible_file = result_file.replace(".tif", "_visible.png")
	
	save_img = Image.fromarray(vote)
	save_img.putpalette([
		100, 100, 100,
		10, 10, 10,
		20, 20, 20,
		30, 30, 30,
		40, 40, 40,
		50, 50, 50,
		60, 60, 60,
		70, 70, 70,
		80, 80, 80,
		90, 90, 90
	])
	save_img.save(result_file)
	dataset_util.save_visible(visible_file, vote)


def calc_result_excel(result_path, test_label_path, each_file_result=True):
	result_files = glob.glob(result_path + "/*.tif")

	for result_file in result_files:
		gt_file = os.path.join(test_label_path, os.path.basename(result_file)[:-4]+"_FGT.tif")

		count_table = calc_accuracy(gt_file, result_file, NUM_CLASSES)
		# print("*" * 20)

		excel_file = os.path.join(result_path, '%s_result.xlsx' % os.path.basename(result_file))
		wb = Workbook()
		ws = wb.active

		for col in range(NUM_CLASSES):
			for row in range(NUM_CLASSES):
				ws.cell(row=row + 1, column=col + 1, value=count_table[row][col])
			#count_table[0][col] = 0

		iou_list = []
		true_count = 0
		class_count = 0
		for class_idx in range(NUM_CLASSES):
			# count_table[0][class_idx] = 0
			tp = count_table[class_idx][class_idx]
			tp_fp = np.sum(count_table, axis=0)[class_idx]
			tp_fn = np.sum(count_table, axis=1)[class_idx]
			if tp_fp + tp_fn == 0:
				iou = 0
			else:
				class_count += 1
				iou = tp / (tp_fp + tp_fn - tp)
				if tp_fn == 0:
					acc = 'NaN'
				else:
					acc = tp / tp_fn
			iou_list.append(iou)
			ws.cell(row=NUM_CLASSES + 1, column=class_idx + 1, value=acc)
			ws.cell(row=NUM_CLASSES + 2, column=class_idx + 1, value=iou)
			true_count += count_table[class_idx][class_idx]

		ws.cell(row=NUM_CLASSES + 1, column=1, value="Acc")
		ws.cell(row=NUM_CLASSES + 2, column=1, value="Iou")
		mIou = np.sum(iou_list) / class_count
		fIou = np.sum(np.multiply(iou_list, np.sum(count_table, axis=1))) / np.sum(count_table)
		accuracy = true_count / np.sum(count_table) #round(true_count / np.sum(count_table), 4)
		ws.cell(row=NUM_CLASSES + 3, column=1, value="mIou")
		ws.cell(row=NUM_CLASSES + 3, column=2, value=mIou)
		ws.cell(row=NUM_CLASSES + 4, column=1, value="fIou")
		ws.cell(row=NUM_CLASSES + 4, column=2, value=fIou)
		ws.cell(row=NUM_CLASSES + 5, column=1, value="accuracy")
		ws.cell(row=NUM_CLASSES + 5, column=2, value=accuracy)

		wb.save(excel_file)

	result_excel = os.path.join(result_path, os.path.basename(result_path) + '.xlsx')
	excel_files = glob.glob(result_path + "/*.xlsx")
	print("result_excel file = ", result_excel)

	wb = Workbook()
	ws = wb.active
	if each_file_result:
		ws_f = wb.create_sheet('file')
		ws_f['A1'].value = "filename"
		ws_f['B1'].value = "Accuracy"
	count_table = np.zeros([NUM_CLASSES, NUM_CLASSES], dtype=np.int)
	for idx, excel_file in enumerate(excel_files):
		all_px = 0
		true_px = 0
		print(excel_file)
		f_wb = load_workbook(excel_file)
		f_ws = f_wb.worksheets[0]
		for i in range(NUM_CLASSES):
			for j in range(NUM_CLASSES):
				count_table[i][j] += int(f_ws.cell(row=i + 1, column=j + 1).value)
				#if i > 0 and j > 0:
				px = f_ws.cell(row=i + 1, column=j + 1).value
				all_px += px
				if i == j:
					true_px += px
		if each_file_result:
			tif_name = os.path.basename(excel_file).replace("_result.xlsx", "")
			ws_f.cell(row=idx + 2, column=1).value = tif_name
			ws_f.cell(row=idx + 2, column=2).value = true_px / all_px
	# print(count_table)

	for col in range(NUM_CLASSES):
		for row in range(NUM_CLASSES):
			ws.cell(row=row + 1, column=col + 1, value=count_table[row][col])
		#count_table[0][col] = 0

	iou_list = []
	true_count = 0
	class_count = 0
	for class_idx in range(NUM_CLASSES):
		tp = count_table[class_idx][class_idx]
		tp_fp = np.sum(count_table, axis=0)[class_idx]
		tp_fn = np.sum(count_table, axis=1)[class_idx]
		if tp_fp + tp_fn == 0:
			iou = 0
			acc = 0
		else:
			class_count += 1
			iou = tp / (tp_fp + tp_fn - tp)
			acc = tp / tp_fn
		iou_list.append(iou)
		ws.cell(row=NUM_CLASSES + 1, column=class_idx + 1, value=acc)
		ws.cell(row=NUM_CLASSES + 2, column=class_idx + 1, value=iou)
		true_count += count_table[class_idx][class_idx]

	mIou = np.sum(iou_list) / class_count
	fIou = np.sum(np.multiply(iou_list, np.sum(count_table, axis=1))) / np.sum(count_table)
	accuracy = round(true_count / np.sum(count_table), 4)
	#ws.cell(row=NUM_CLASSES + 1, column=1, value="Acc")
	#ws.cell(row=NUM_CLASSES + 2, column=1, value="Iou")
	ws.cell(row=NUM_CLASSES + 3, column=1, value="mIou")
	ws.cell(row=NUM_CLASSES + 4, column=1, value="fIou")
	ws.cell(row=NUM_CLASSES + 5, column=1, value="accuracy")
	ws.cell(row=NUM_CLASSES + 3, column=2, value=mIou)
	ws.cell(row=NUM_CLASSES + 4, column=2, value=fIou)
	ws.cell(row=NUM_CLASSES + 5, column=2, value=accuracy)

	print('Iou = ', iou_list)
	print(f'Result (mIou, fIou, accuracy) : {mIou}, {fIou}, {accuracy}')

	wb.save(result_excel)

	return accuracy

def inference_mindAT(sess, predictions, placeholder, image_ori, gpu_id=None):
	'''

	:param weight_path: path of weight
	:param test_image: test_image [numpy array from imageio.imread]
	:param gpu_id:
	:return vote: numpy array of image
	'''

	# os.makedirs(result_path, exist_ok=True)
	if gpu_id != None:
		os.environ["CUDA_VISIBLE_DEVICES"] = str(gpu_id)

	# x = tf.placeholder(tf.float32, [1, PATCH_SIZE, PATCH_SIZE, DEPTH])

	# predictions = deeplab_model.deeplabv3_plus_model_fn(
	# 	x,
	# 	None,
	# 	tf.estimator.ModeKeys.PREDICT,
	# 	params={
	# 		'output_stride': output_stride,
	# 		'batch_size': 1,  # Batch size must be 1 because the images' size may differ
	# 		'base_architecture': base_architecture,
	# 		'pre_trained_model': None,
	# 		'batch_norm_decay': None,
	# 		'num_classes': NUM_CLASSES,
	# 		'freeze_batch_norm': True,
	# 		'num_channels': DEPTH
	# 	}
	# ).predictions

	# saver = tf.train.Saver()
	# ckpt = tf.train.get_checkpoint_state(weight_path)

	# os.makedirs(result_path, exist_ok=True)
	# files = os.listdir(test_image_path)
	# for filename in files:
	# image_path = os.path.join(test_image_path, filename)

	row_ori = image_ori.shape[0]
	col_ori = image_ori.shape[1]
	depth_ori = image_ori.shape[2]

	vote = np.zeros((row_ori, col_ori, NUM_CLASSES))

	if depth_ori < DEPTH:
		zero_ch = np.zeros((image_ori.shape[0], image_ori.shape[1], DEPTH - depth_ori))
		image = np.dstack((image_ori, zero_ch))
		# image = resize(image, (image_3ch.shape[1], image_3ch.shape[0], 5))
		image = np.uint8(image)
	else:
		image = image_ori[..., 0:DEPTH]
	# print("0 image, ", image_ori.shape)
	if DEPTH == 3:
		rgb_mean = np.asarray(
			[preprocessing._R_MEAN, preprocessing._G_MEAN, preprocessing._B_MEAN] * row_ori * col_ori)
	elif DEPTH == 4:
		rgb_mean = np.asarray(
			[preprocessing._R_MEAN, preprocessing._G_MEAN, preprocessing._B_MEAN, 0.0] * row_ori * col_ori)
	else:
		rgb_mean = np.asarray(
			[preprocessing._R_MEAN, preprocessing._G_MEAN, preprocessing._B_MEAN, 0.0, 0.0] * row_ori * col_ori)
	rgb_mean = np.reshape(rgb_mean, [row_ori, col_ori, DEPTH])
	image = image - rgb_mean

	row = image.shape[0]
	col = image.shape[1]
	# vote_each = np.zeros((row, col, NUM_CLASSES))

# with tf.Session() as sess:
	# saver.restore(sess, ckpt.model_checkpoint_path)

	row_list = list(range(0, row - PATCH_SIZE, int(PATCH_SIZE / 2)))
	row_list.append(row - PATCH_SIZE)
	col_list = list(range(0, col - PATCH_SIZE, int(PATCH_SIZE / 2)))
	col_list.append(col - PATCH_SIZE)

	for r in row_list:
		for c in col_list:
			r_off = 0
			c_off = 0

			sub_image = image[r + r_off:r + PATCH_SIZE + r_off, c + c_off:c + PATCH_SIZE + c_off, :]
			sub_image = np.reshape(sub_image, [1, PATCH_SIZE, PATCH_SIZE, DEPTH])

			preds = sess.run(predictions, feed_dict={placeholder: sub_image})
			pred_prob = preds['probabilities']
			crop_result = np.reshape(pred_prob, (PATCH_SIZE, PATCH_SIZE, NUM_CLASSES))
			vote[r + r_off:r + PATCH_SIZE + r_off, c + c_off:c + PATCH_SIZE + c_off] += crop_result
# sess.close()

	return vote

def inference(weight_path, test_image_path, result_path, gpu_id):
	print("inference, weight_path = {}, test_image_path = {}, result_path = {}".format(weight_path, test_image_path, result_path))
	os.makedirs(result_path, exist_ok=True)
	if gpu_id != None:
		os.environ["CUDA_VISIBLE_DEVICES"] = str(gpu_id)

	x = tf.placeholder(tf.float32, [1, PATCH_SIZE, PATCH_SIZE, DEPTH])

	predictions = deeplab_model.deeplabv3_plus_model_fn(
		x,
		None,
		tf.estimator.ModeKeys.PREDICT,
		params={
			'output_stride': output_stride,
			'batch_size': 1,  # Batch size must be 1 because the images' size may differ
			'base_architecture': base_architecture,
			'pre_trained_model': None,
			'batch_norm_decay': None,
			'num_classes': NUM_CLASSES,
			'freeze_batch_norm': True,
			'num_channels': DEPTH
		}
	).predictions

	saver = tf.train.Saver()
	ckpt = tf.train.get_checkpoint_state(weight_path)

	os.makedirs(result_path, exist_ok=True)
	files = os.listdir(test_image_path)
	for filename in files:
		image_path = os.path.join(test_image_path, filename)
		image_ori = dataset_util.load_image(image_path, read_label=False, bit16=False)

		row_ori = image_ori.shape[0]
		col_ori = image_ori.shape[1]
		depth_ori = image_ori.shape[2]

		print(filename, row_ori, col_ori, depth_ori)

		vote = np.zeros((row_ori, col_ori, NUM_CLASSES))

		if depth_ori < DEPTH:
			zero_ch = np.zeros((image_ori.shape[0], image_ori.shape[1], DEPTH - depth_ori))
			image = np.dstack((image_ori, zero_ch))
			# image = resize(image, (image_3ch.shape[1], image_3ch.shape[0], 5))
			image = np.uint8(image)
		else:
			image = image_ori[..., 0:DEPTH]
		# print("0 image, ", image_ori.shape)
		if DEPTH == 3:
			rgb_mean = np.asarray(
				[preprocessing._R_MEAN, preprocessing._G_MEAN, preprocessing._B_MEAN] * row_ori * col_ori)
		elif DEPTH == 4:
			rgb_mean = np.asarray(
				[preprocessing._R_MEAN, preprocessing._G_MEAN, preprocessing._B_MEAN, 0.0] * row_ori * col_ori)
		else:
			rgb_mean = np.asarray(
				[preprocessing._R_MEAN, preprocessing._G_MEAN, preprocessing._B_MEAN, 0.0, 0.0] * row_ori * col_ori)
		rgb_mean = np.reshape(rgb_mean, [row_ori, col_ori, DEPTH])
		image = image - rgb_mean

		row = image.shape[0]
		col = image.shape[1]
		# vote_each = np.zeros((row, col, NUM_CLASSES))

		with tf.Session() as sess:
			saver.restore(sess, ckpt.model_checkpoint_path)

			row_list = list(range(0, row - PATCH_SIZE, int(PATCH_SIZE / 2)))
			row_list.append(row - PATCH_SIZE)
			col_list = list(range(0, col - PATCH_SIZE, int(PATCH_SIZE / 2)))
			col_list.append(col - PATCH_SIZE)

			for r in row_list:
				for c in col_list:
					r_off = 0
					c_off = 0

					sub_image = image[r + r_off:r + PATCH_SIZE + r_off, c + c_off:c + PATCH_SIZE + c_off, :]
					sub_image = np.reshape(sub_image, [1, PATCH_SIZE, PATCH_SIZE, DEPTH])

					preds = sess.run(predictions, feed_dict={x: sub_image})
					pred_prob = preds['probabilities']
					crop_result = np.reshape(pred_prob, (PATCH_SIZE, PATCH_SIZE, NUM_CLASSES))
					vote[r + r_off:r + PATCH_SIZE + r_off, c + c_off:c + PATCH_SIZE + c_off] += crop_result

		save_result(filename, vote, result_path)
	sess.close()

	
def run(weight_path, test_image_path, test_label_path, result_path, gpu_id):
	print("evaluate, weight_path = {}, test_image_path = {}, test_label_path = {}, result_path = {}"
	      .format(weight_path, test_image_path, test_label_path, result_path))
	
	inference(weight_path, test_image_path, result_path, gpu_id)

	old_excels = glob.glob(result_path + "/*.xlsx")
	for old_excel in old_excels:
		os.remove(old_excel)

	accuracy = calc_result_excel(result_path, test_label_path)

	return accuracy


def ensemble(model_list, test_image_path, test_label_path, result_path):
	offset_list = [[0, 0], [0, 1], [1, 0], [1, 1], [2, 0], [0, 2], [3, 0], [0, 3], [2, 1], [1, 2], [4, 0], [0, 4],
	               [3, 1], [2, 2], [1, 3]]

	os.makedirs(result_path, exist_ok=True)
	files = os.listdir(test_image_path)
	for filename in files:
		image_path = os.path.join(test_image_path, filename)
		label_path = os.path.join(test_label_path, filename)
		image_ori, label = dataset_util.load_image(image_path, label_path, read_label=True, bit16=False)

		row_ori = image_ori.shape[0]
		col_ori = image_ori.shape[1]

		print(filename, row_ori, col_ori)

		vote = np.zeros((row_ori, col_ori, NUM_CLASSES))

		for model_idx, model_path in enumerate(model_list):
			image = image_ori[..., 0:DEPTH]
			# print("0 image, ", image_ori.shape)
			if DEPTH == 3:
				rgb_mean = np.asarray(
					[preprocessing._R_MEAN, preprocessing._G_MEAN, preprocessing._B_MEAN] * row_ori * col_ori)
			elif DEPTH == 4:
				rgb_mean = np.asarray(
					[preprocessing._R_MEAN, preprocessing._G_MEAN, preprocessing._B_MEAN, 0.0] * row_ori * col_ori)
			else:
				rgb_mean = np.asarray(
					[preprocessing._R_MEAN, preprocessing._G_MEAN, preprocessing._B_MEAN, 0.0, 0.0] * row_ori * col_ori)
			rgb_mean = np.reshape(rgb_mean, [row_ori, col_ori, DEPTH])
			image = image - rgb_mean

			rot = model_idx % 4

			if rot % 4 != 0:
				image = scipy.ndimage.rotate(image, rot * 90, order=0)
			row = image.shape[0]
			col = image.shape[1]
			vote_each = np.zeros((row, col, NUM_CLASSES))

			tf.reset_default_graph()
			x = tf.placeholder(tf.float32, [1, PATCH_SIZE, PATCH_SIZE, DEPTH])

			predictions = deeplab_model.deeplabv3_plus_model_fn(
				x,
				None,
				tf.estimator.ModeKeys.PREDICT,
				params={
					'output_stride': output_stride,
					'batch_size': 1,  # Batch size must be 1 because the images' size may differ
					'base_architecture': base_architecture,
					'pre_trained_model': None,
					'batch_norm_decay': None,
					'num_classes': NUM_CLASSES,
					'freeze_batch_norm': True,
					'num_channels': DEPTH
				}
			).predictions

			# Manually load the latest checkpoint
			saver = tf.train.Saver()

			with tf.Session() as sess:
				ckpt = tf.train.get_checkpoint_state(model_path)
				saver.restore(sess, ckpt.model_checkpoint_path)

				offset_idx = model_idx % (len(offset_list))
				r_start = int(offset_list[offset_idx][0] * PATCH_SIZE / 8)
				c_start = int(offset_list[offset_idx][1] * PATCH_SIZE / 8)

				row_list = list(range(r_start, row - PATCH_SIZE, int(PATCH_SIZE / 2)))
				if row-PATCH_SIZE not in row_list:
					row_list.append(row - PATCH_SIZE)
				col_list = list(range(c_start, col - PATCH_SIZE, int(PATCH_SIZE / 2)))
				if col - PATCH_SIZE not in col_list:
					col_list.append(col - PATCH_SIZE)

				for r in row_list:
					for c in col_list:
						sub_image = image[r : r + PATCH_SIZE , c : c + PATCH_SIZE , :]
						sub_image = np.reshape(sub_image, [1, PATCH_SIZE, PATCH_SIZE, DEPTH])

						preds = sess.run(predictions, feed_dict={x: sub_image})
						pred_prob = preds['probabilities']
						crop_result = np.reshape(pred_prob, (PATCH_SIZE, PATCH_SIZE, NUM_CLASSES))
						vote_each[r : r + PATCH_SIZE , c : c + PATCH_SIZE ] += crop_result

			if rot % 4 != 0:
				vote_each = scipy.ndimage.rotate(vote_each, -rot * 90, order=0)
			vote += vote_each

			sess.close()
		save_result(filename, vote, result_path)

	accuracy = calc_result_excel(result_path, test_label_path)

	return accuracy
